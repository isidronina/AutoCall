from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime, date
import os

app = Flask(__name__)
print(app.url_map)
EXCEL_PATH = 'Lista_de_Chamada.xlsx'
SHEET = 'Presenças'
print("Excel em:", os.path.abspath(EXCEL_PATH))
def carregar_dados():
    # 1. Caminho absoluto
    abs_path = os.path.abspath(EXCEL_PATH)
    print(f"[DEBUG] Lendo Excel em: {abs_path}")

    # 2. Hora de modificação do arquivo
    mtime = os.path.getmtime(EXCEL_PATH)
    print(f"[DEBUG] Última modificação: {datetime.fromtimestamp(mtime)}")

    # 3. Leitura
    df = pd.read_excel(
        EXCEL_PATH,
        sheet_name=SHEET,
        engine='openpyxl',
        dtype={'TURMA': str, 'RA': str}
    )
    df['TURMA'] = df['TURMA'].str.strip()
    df['RA']    = df['RA'].str.strip()

    # 4. Mostra as 5 primeiras linhas
    print("[DEBUG] DataFrame carregado:\n", df.head(), "\n")

    return df

def salvar_dados(df):
    df.to_excel(EXCEL_PATH, sheet_name=SHEET, index=False, engine='openpyxl')



def _sanitize_sheet_name(name: str) -> str:
    """Excel limita a 31 chars e proíbe []:*?/\\; também não pode vazio."""
    invalid = set('[]:*?/\\')
    safe = ''.join('_' if ch in invalid else ch for ch in (name or '').strip())
    safe = safe or 'Turma'
    return safe[:31]


def registrar_log(turma, nome, ra, data, hora, presenca):
    wb = load_workbook(EXCEL_PATH)

    # 1) Log geral (aba "Registros")
    if 'Registros' not in wb.sheetnames:
        ws_reg = wb.create_sheet('Registros')
        ws_reg.append(['TURMA','NOME','RA','DATA','HORA','PRESENCA'])
    else:
        ws_reg = wb['Registros']
    ws_reg.append([turma, nome, ra, data, hora, presenca])

    # 2) Log por turma (uma aba para cada turma)
    sheet_name = _sanitize_sheet_name(str(turma))
    if sheet_name not in wb.sheetnames:
        ws_turma = wb.create_sheet(sheet_name)
        ws_turma.append(['TURMA','NOME','RA','DATA','HORA','PRESENCA'])
    else:
        ws_turma = wb[sheet_name]
        # se a aba existir mas estiver vazia, garante cabeçalho
        if ws_turma.max_row == 1 and all(c.value is None for c in ws_turma[1]):
            ws_turma.append(['TURMA','NOME','RA','DATA','HORA','PRESENCA'])

    ws_turma.append([turma, nome, ra, data, hora, presenca])

    wb.save(EXCEL_PATH)

@app.route('/chamada')
def chamada():
    turma = request.args.get('turma', '').strip()

    # pega a data da URL; se não vier, usa hoje
    data_str = (request.args.get('data') or '').strip()
    try:
        data_ref = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else date.today()
    except ValueError:
        data_ref = date.today()
    data_iso = data_ref.isoformat()

    df = carregar_dados()
    df_turma = df[df['TURMA'] == turma].copy()

    df_turma.rename(columns={
        'NOME': 'nome',
        'RA': 'ra',
        'HORARIO': 'hora',
        'PRESENÇA': 'presenca'
    }, inplace=True)

    df_turma['presenca'] = 'Ausente'
    df_turma['hora'] = ''

    # lê o log e marca quem passou NA DATA ESCOLHIDA
    wb = load_workbook(EXCEL_PATH, data_only=True)
    if 'Registros' in wb.sheetnames:
        df_log = pd.read_excel(EXCEL_PATH, sheet_name='Registros', engine='openpyxl', dtype=str)
        df_log = df_log[(df_log['TURMA'] == turma) & (df_log['DATA'] == data_iso)]
        for _, row in df_log.iterrows():
            mask = df_turma['ra'] == row['RA'].strip()
            df_turma.loc[mask, 'presenca'] = row['PRESENCA']
            df_turma.loc[mask, 'hora'] = row['HORA']

    df_turma.insert(0, 'id', range(1, len(df_turma)+1))
    dados = df_turma[['id','nome','ra','hora','presenca']].to_dict(orient='records')

    # manda para o template a data que estamos filtrando
    return render_template('chamada.html', turma=turma, dados=dados, hoje=data_iso)

@app.route('/')
def index():
    df = carregar_dados()
    turmas = sorted(df['TURMA'].dropna().unique())
    return render_template('index.html', turmas=turmas)


@app.route('/scan', methods=['POST'])
def scan():
    data = request.get_json(silent=True) or {}
    ra_scanned = (data.get('ra') or '').strip()
    turma = (data.get('turma') or '').strip()
    data_payload = (data.get('data') or '').strip()  # <- data vinda do front

    if not ra_scanned:
        return jsonify({'error': 'RA ausente'}), 400

    df = carregar_dados()
    if turma:
        df_match = df[(df['TURMA'] == turma) & (df['RA'] == ra_scanned)]
        if df_match.empty:
            return jsonify({'error': 'RA não encontrado na turma informada'}), 404
    else:
        df_match = df[df['RA'] == ra_scanned]
        turmas = df_match['TURMA'].dropna().unique().tolist()
        if len(turmas) == 0:
            return jsonify({'error': 'RA não encontrado'}), 404
        if len(turmas) > 1:
            return jsonify({'error': f'RA encontrado em múltiplas turmas: {", ".join(turmas)}'}), 409
        turma = turmas[0]
        df_match = df_match[df_match['TURMA'] == turma]

    row = df_match.iloc[0]
    nome = row['NOME']

    now = datetime.now()
    # se o front enviar uma data válida, usa ela; senão, hoje
    try:
        data_hoje = datetime.strptime(data_payload, '%Y-%m-%d').date().isoformat() if data_payload else now.date().isoformat()
    except ValueError:
        data_hoje = now.date().isoformat()
    hora = now.strftime('%H:%M:%S')

    registrar_log(turma, nome, ra_scanned, data_hoje, hora, 'Presente')

    return jsonify({'nome': nome, 'ra': ra_scanned, 'hora': hora, 'presenca': 'Presente', 'turma': turma})



if __name__ == '__main__':
    print(app.url_map)    # vê todas as rotas aqui
    app.run(debug=True)

